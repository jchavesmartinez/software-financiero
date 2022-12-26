# %%
# Esto es para cargar todas las librerias

from ctypes.wintypes import SIZE
import PySimpleGUI as sg
import pandas as pd
from datetime import date
from datetime import datetime
from win32com import client

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string
from openpyxl.styles import PatternFill, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
import numpy as np
import sys
import os

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import io
import PIL.Image
import shutil




# %%
def VentanaInventarios():
    # Esto es para generar una parte del layout
    path= os.getcwd()+'\\'

    InventarioHistorico = pd.read_csv(path+'InventarioHistorico.csv', encoding='latin-1')

    inventarioSA = InventarioHistorico.values.tolist()
    inventarioSA_heads = ['Fecha','Código', 'Producto', 'Precio Unitario', 'Costo unitario', 'Movimiento']
    inventarioSA_width = [9,12, 40, 11,0,8]
            
    InventarioSuperiorA= [

            [sg.Text("Fecha  ", size=(10,1)),sg.Text("Codigo de barra #",size=(20,1)),sg.Text("Producto")],
            [sg.Input(date.today().strftime("%b-%d-%Y"), size=(12,1), key="-FECHA-", disabled=True),sg.Input("", size=(22,1),key="-CODIGO-", enable_events=True),sg.Input("", size=(61,None),key="-PRODUCTO-"),sg.Button("", size=(3,1),key="-AGREGARFOTO-")],
            [sg.Text("Cantidad"),sg.Input("", size=(7,None),key="-CANTIDAD-"),sg.Text("Costo Unitario"),sg.Input("", size=(13,None), key="-COSTO-"),sg.Text("Precio de venta unitario"),sg.Input("", size=(13,None), key="-PRECIOVENTA-"),sg.Radio('Contado', "RADIO1",key="-CONTADO-", default=True),sg.Radio('Crédito', "RADIO1",key="-CREDITO-")],
            [sg.Text("Proveedor"),sg.Input("",size=(51,2), key="-PROVEEDOR-", enable_events=True),sg.Text("Correo electronico "),sg.Input("", size=(24,2), key="-EMAILPROVEEDOR-", enable_events=True)],
            [sg.Button("Registrar entrada", size=(44,2),key="-REGISTRARENTRADA-"),sg.Button("Eliminar entrada", size=(45,2),key="-ELIMINARENTRADA-")],
            [sg.Text("Buscar por nombre o codigo"),sg.Input("",expand_x=True, key="-FILTRARINVENTARIO-", enable_events=True),sg.Button("Disponible", size=(9,1),key="-VISUALIZARINVENTARIO-")],
            [sg.Table(values=inventarioSA, enable_events=True, headings=inventarioSA_heads,max_col_width=65,col_widths=inventarioSA_width,auto_size_columns=False,justification='left',num_rows=8,background_color="White", row_height=22,text_color="black", key="-INVENTARIOHISTORICO-")],


    ]
    

    #url = 'https://raw.githubusercontent.com/jchavesmartinez/software-financiero/main/cuentas%20-%20Copy.csv'
    totalcuentas = pd.read_csv(path+'cuentas - Copy.csv')
    totalcuentas=totalcuentas.drop(['Nivel', 'Primer Nivel'], axis=1)

    registrodataSB=totalcuentas
    registrodataSB_heads = ['Cedula', 'Cliente']
    registrodataSB=registrodataSB.values.tolist()
    registrodataSB_width = [12, 50]

    InventarioSuperiorB= [
        [sg.Text("Nombre cliente"),sg.Input("",size=(36,2), key="-NOMBRECLIENTE-", enable_events=True),sg.Text("Cedula cliente"),sg.Input("",size=(15,2), key="-FILTRARCUENTAS-", enable_events=True)],
        [sg.Text("Correo electronico"),sg.Input("",size=(36,2), key="-CORREOCLIENTE-", enable_events=True),sg.Text("Celular "),sg.Input("",size=(18,2), key="-CELULARCLIENTE-", enable_events=True)],
        [sg.Table(values=registrodataSB, headings=registrodataSB_heads,max_col_width=65,col_widths=registrodataSB_width,auto_size_columns=False,justification='left',num_rows=3,background_color="White", row_height=25,text_color="black", key="-CUENTAS-", enable_events=True,select_mode=sg.TABLE_SELECT_MODE_BROWSE)],
        [sg.Text("Codigo de barras"),sg.Input("",size=(18,2), key="-CODIGOSALIDA-", enable_events=True),sg.Text("Producto"),sg.Input("",size=(36,2), key="-PRODUCTOSALIDA-", enable_events=True)],
        [sg.Text("Cantidad"),sg.Input("", size=(5,None),key="-CANTIDADSALIDA-"),sg.Text("Precio de venta"),sg.Input("", size=(13,None), key="-PRECIOVENTASALIDA-"),sg.Radio('Contado', "RADIO2",key="-CONTADO-", default=True),sg.Radio('Crédito', "RADIO2",key="-CREDITO-"),sg.Radio('Devolucion', "RADIO2",key="-DEVOLUCION-")],
        [sg.Text("Clave factura"),sg.Input("",size=(69,2), key="-CLAVEFACTURA-", enable_events=True)],
        [sg.Table(values=registrodataSB, headings=registrodataSB_heads,max_col_width=65,col_widths=registrodataSB_width,auto_size_columns=False,justification='left',num_rows=3,background_color="White", row_height=29,text_color="black", key="-CUENTAS-", enable_events=True,select_mode=sg.TABLE_SELECT_MODE_BROWSE)],
        [sg.Button("Registrar salida", size=(71,2), key="-REGISTROSALIDA-")]

    ]

    # Esta funcion cambia el titulo de las tablas

    def update_title(table, headings):
        for cid, text in zip(inventarioSA_heads, headings):
            table.heading(cid, text=text)

    def update_width(table,anchos):
        for cid, width in zip(inventarioSA_heads, anchos):    # Set width for each column
            table.column(cid, width=width)

    # funcion para cambiar tamaño de la imagen

    def convToBytes(image, resize=(200,200)):
        
        img = image.copy()	
        cur_width, cur_height = img.size
        if resize:
            new_width, new_height = resize
            scale = min(new_height/cur_height, new_width/cur_width)
            img = img.resize((int(cur_width*scale), int(cur_height*scale)), PIL.Image.ANTIALIAS)	
        ImgBytes = io.BytesIO()
        img.save(ImgBytes, format="PNG")
        del img
        return ImgBytes.getvalue()

    # esto es para generar los graficos

    Graficos= [
        [sg.Image(path+'noimage.png',size=(200, 200),key='-PRODUCTOFOTO-'),sg.Canvas(key='-DEMANDA-',pad=(0,0)),sg.Canvas(key='-CYBGRAFICO-',pad=(0,0))],

    ]

    # Esto consolida todas las partes de la intefaz en un  solo layout

    menucompleto = [
        [sg.Frame("", [[sg.Frame("ENTRADAS", InventarioSuperiorA,border_width=0),sg.Frame("SALIDAS", InventarioSuperiorB,border_width=0)]],border_width=0, pad=(0,0))],
        [sg.Frame("", [[sg.Frame("VISUALES", Graficos,border_width=0)]],border_width=0, pad=(0,0))],
    ]
    
    menu_def = [['Menu', ['Contabilidad','Manejo de Inventarios','Facturación', 'Manejo de Planilla', 'Analísis Financiero', 'Cuentas por pagar y cobrar']],['Info', ['Manual de Uso','Preguntas y Respuestas','Reportar Error']],['Acerca de', ['Quienes somos','Contacto']]]
    layout = [
        [sg.Menu(menu_def, tearoff=False, pad=(200, 1))],
        [sg.Frame("", menucompleto,border_width=0,expand_y=True)]
    ]
    
    right_click_menu = ['Unused', ['Forecasting App', '!&Click', '&Menu', 'E&xit', 'Properties']]
    window= sg.Window("Manejo de inventarios",layout,finalize=True,icon=path+'favicon2.ico',right_click_menu=right_click_menu)

    # Obtención del canvas
    from turtle import color


    canvas = window['-CYBGRAFICO-'].TKCanvas
    figure = Figure(facecolor='#64778D',figsize=(3,3))
    axes = figure.add_subplot()
    figure_canvas_agg = FigureCanvasTkAgg(figure, canvas)
    figure_canvas_agg.draw()
    figure_canvas_agg.get_tk_widget().pack(side='bottom', fill='both', expand=1)

    canvas2 = window['-DEMANDA-'].TKCanvas
    figure2 = Figure(facecolor='#64778D',figsize=(9,3))
    axes2 = figure2.add_subplot()
    figure_canvas2_agg = FigureCanvasTkAgg(figure2, canvas2)
    figure_canvas2_agg.draw()
    figure_canvas2_agg.get_tk_widget().pack(side='bottom', fill='both', expand=1)

    def costobeneificio():

        try:
            movimiento=[]

            for x in range(0, len(inventarioSA)):
                movimiento.append(inventarioSA[x][5])
                neto=sum(map(float, movimiento))

            valor=[]

            for x in range(0, len(inventarioSA)):
                valor.append(inventarioSA[x][6])
                valortotal=sum(map(float, valor))

            costo=[]

            for x in range(0, len(inventarioSA)):
                costo.append(inventarioSA[x][7])
                costototal=sum(map(float, costo))

            labels = ['G1']
            costoplot = [costototal/valortotal*100]
            valorplot = [((valortotal-costototal)/valortotal)*100]
            width = 0.5

            axes.cla()
            axes.bar(labels, costoplot, width, label='% Costo',color=['#1A263B', 'white'])
            axes.bar(labels, valorplot, width, bottom=costoplot,label='% Margen ganancia')
            axes.set_facecolor("#64778D") 
            axes.set_title('Margen de Ganancia Inventario',color="white")
            axes.legend(loc='center left', bbox_to_anchor=(0.40,0.8))
            axes.spines['top'].set_visible(False)
            axes.spines['right'].set_visible(False)
            axes.spines['bottom'].set_color('white')
            axes.spines['left'].set_color('white')
            axes.tick_params(axis='x', colors='white')
            axes.tick_params(axis='y', colors='white')
        except:
            sg.Popup('No existe registro')


    def demandagrafico():

        InventarioTotal1 = pd.DataFrame(inventarioSA, columns=['Fecha','Código', 'Producto', 'Precio Unitario', 'Costo unitario', 'Movimiento', 'Precio Total', 'Costo Total', 'Tipo'])
        InventarioTotal1=InventarioTotal1.drop(['Precio Unitario','Costo unitario','Costo Total','Precio Total','Tipo'], axis=1)
        InventarioTotal1['Movimiento'] = InventarioTotal1['Movimiento'].fillna(0).astype(int)
        InventarioTotal1['InvetarioNeto'] = InventarioTotal1['Movimiento'].cumsum()

        InventarioTotal1=InventarioTotal1.values.tolist()

        InventarioTotal2 = pd.DataFrame(inventarioSA, columns=['Fecha','Código', 'Producto', 'Precio Unitario', 'Costo unitario', 'Movimiento', 'Precio Total', 'Costo Total', 'Tipo'])
        InventarioTotal2=InventarioTotal2.drop(['Precio Unitario','Costo unitario','Costo Total','Precio Total','Tipo','Fecha'], axis=1)
        InventarioTotal2['Movimiento'] = InventarioTotal2['Movimiento'].fillna(0).astype(int)
        InventarioTotal2['Inventario Prom']=InventarioTotal2['Movimiento'].cumsum()
        InventarioTotal2=InventarioTotal2.drop(['Movimiento'], axis=1)
        InventarioTotal2=InventarioTotal2.mean()

        InventarioTotal3 = pd.DataFrame(inventarioSA, columns=['Fecha','Código', 'Producto', 'Precio Unitario', 'Costo unitario', 'Movimiento', 'Precio Total', 'Costo Total', 'Tipo'])
        InventarioTotal3=InventarioTotal3[InventarioTotal3["Tipo"] == 'Salida'] 
        InventarioTotal3 = abs(InventarioTotal3["Movimiento"].mean())

        InventarioNeto=[]
        sucesion=[]
        fecha=[]
        prom=[]
        demanda=[]

        for x in range(0, len(InventarioTotal1)):
            InventarioNeto.append(InventarioTotal1[x][4])
            sucesion.append([x])
            fecha.append(InventarioTotal1[x][0])
            prom.append(InventarioTotal2)
            demanda.append(InventarioTotal3)


        axes2.cla()
        axes2.plot(sucesion, InventarioNeto, label = "Inventario Neto", color="#D87D2D")
        axes2.plot(sucesion, prom, label = "Inventario Promedio", linestyle='dashed', color="#1A263B")
        axes2.plot(sucesion, demanda, label = "Demanda Promedio", linestyle='dashdot', color="#E3ED68")
        axes2.set_title('Comportamiento del inventario historico',color="white", loc='right')
        axes2.set_xticks(range(0,len(fecha)))
        axes2.set_xticklabels(fecha)
        axes2.tick_params(axis='x', colors='white')
        axes2.tick_params(axis='y', colors='white')
        axes2.legend(loc='center left', bbox_to_anchor=(0.0,1))
        axes2.set_facecolor("#64778D")
        axes2.spines['top'].set_visible(False)
        axes2.spines['right'].set_visible(False)
        axes2.spines['bottom'].set_color('white')
        axes2.spines['left'].set_color('white')
        figure2.autofmt_xdate()

    # Esto es el codigo para ejecutar la interfaz

    from faulthandler import disable

    if len(inventarioSA)!=0: 
        costobeneificio()
        figure_canvas_agg.draw()

        demandagrafico()
        figure_canvas2_agg.draw()

        foldername="vacio"


    while True:
        
        event, values = window.read()
        
        if event == sg.WIN_CLOSED:
            break

        if event == '-VISUALIZARINVENTARIO-' and window['-VISUALIZARINVENTARIO-'].get_text()=="Movimientos":
            window['-VISUALIZARINVENTARIO-'].update("Disponible")

            update_title(window['-INVENTARIOHISTORICO-'].Widget, ['Fecha','Código', 'Producto', 'Precio Unitario', 'Costo unitario', 'Movimiento'])
            update_width(window['-INVENTARIOHISTORICO-'].Widget, [83,103, 360, 100,0,80])
            window['-INVENTARIOHISTORICO-'].update(inventarioSA)
            window['-ELIMINARENTRADA-'].update(disabled=False)

        else:
            if event == '-VISUALIZARINVENTARIO-' and window['-VISUALIZARINVENTARIO-'].get_text()=="Disponible":
                window['-VISUALIZARINVENTARIO-'].update("Movimientos")

                InventarioTotal = pd.read_csv(path+'InventarioHistorico.csv', encoding='latin-1')
                InventarioDisponiblePositive = InventarioTotal[InventarioTotal["Tipo"] == 'Entrada'] 
                InventarioDisponiblePositive = InventarioDisponiblePositive.groupby(['Código','Producto']).agg(Movimiento=('Movimiento','sum'), Precio= ('Precio Total','sum'), Costo= ('Costo Total','sum'),InventarioPromedio=('Costo Total','mean'))

                InventarioDisponibleNegative = InventarioTotal[InventarioTotal["Tipo"] == 'Salida'] 
                InventarioDisponibleNegative = InventarioDisponibleNegative.groupby(['Código','Producto']).agg(Movimiento=('Movimiento','sum'), Precio= ('Precio Total','sum'), Costo= ('Costo Total','sum'),InventarioPromedio=('Costo Total','mean'))

                InventarioFinal = pd.concat([InventarioDisponibleNegative, InventarioDisponiblePositive]).groupby(['Código','Producto'])['Movimiento','Precio','Costo'].sum()

                InventarioPromedio=InventarioTotal.drop(['Movimiento','Precio Unitario',"Costo unitario","Fecha","Tipo","Precio Total"], axis=1)
                InventarioPromedio['Inventario Promedio']=InventarioPromedio.groupby(['Código','Producto'])['Costo Total'].cumsum()
                InventarioPromedio=InventarioPromedio.drop(['Costo Total'], axis=1)
                InventarioPromedio=InventarioPromedio.groupby(['Código','Producto']).mean()

                InventarioPromedioUD=InventarioTotal.drop(['Costo Total','Precio Unitario',"Costo unitario","Fecha","Tipo","Precio Total"], axis=1)
                InventarioPromedioUD['Inventario PromedioUD']=InventarioPromedioUD.groupby(['Código','Producto'])['Movimiento'].cumsum()
                InventarioPromedioUD=InventarioPromedioUD.drop(['Movimiento'], axis=1)
                InventarioPromedioUD=InventarioPromedioUD.groupby(['Código','Producto']).mean()


                InventarioFinal['Inventario Promedio1'] = InventarioPromedio
                InventarioFinal['Inventario Promedio'] = InventarioPromedioUD


                InventarioFinal["Rotacion Inventarios"] = 365/(abs(InventarioDisponibleNegative["Costo"])/InventarioFinal['Inventario Promedio1'])

                InventarioFinal=InventarioFinal.drop(['Precio','Inventario Promedio1'], axis=1)
                InventarioFinal=InventarioFinal.reset_index()
                InventarioFinal['Costo'] = InventarioFinal['Rotacion Inventarios'].fillna(0).astype(int)
                InventarioFinal['Rotacion Inventarios'] = InventarioFinal['Rotacion Inventarios'].fillna(0).astype(int)
                InventarioFinal['Inventario Promedio'] = InventarioFinal['Inventario Promedio'].fillna(0).astype(int)
                InventarioFinal=InventarioFinal.values.tolist()


                update_title(window['-INVENTARIOHISTORICO-'].Widget, ['Código', 'Producto', 'En stock', 'Costo Total',"Stock Prom.",'Rotación (días)'])
                update_width(window['-INVENTARIOHISTORICO-'].Widget, [123, 300, 60, 72, 76, 92])
                window['-INVENTARIOHISTORICO-'].update(InventarioFinal)

                window['-ELIMINARENTRADA-'].update(disabled=True)

                
            
        if event == '-FILTRARINVENTARIO-' and values['-FILTRARINVENTARIO-']: 

            InventarioHistorico = pd.read_csv(path+'InventarioHistorico.csv', encoding='latin-1')
            filtrarcuentas=InventarioHistorico.applymap(str)
            filtro=filtrarcuentas[filtrarcuentas.stack().str.contains(values['-FILTRARINVENTARIO-']).any(level=0)].values.tolist()
            window['-INVENTARIOHISTORICO-'].update(filtro)
            inventarioSA=filtro

            window['-REGISTRARENTRADA-'].update(disabled=True)
            
            try:
                window['-PRODUCTOFOTO-'].update(convToBytes(PIL.Image.open(path+inventarioSA[0][1]+".png")))
            except:
                sg.Popup('No existe foto para este producto')

            window['-VISUALIZARINVENTARIO-'].update("Disponible")
            
            costobeneificio()
            figure_canvas_agg.draw()

            demandagrafico()
            figure_canvas2_agg.draw()
            
        if event == '-FILTRARINVENTARIO-' and values['-FILTRARINVENTARIO-']=="":
            InventarioHistorico = pd.read_csv(path+'InventarioHistorico.csv', encoding='latin-1')
            window['-INVENTARIOHISTORICO-'].update(InventarioHistorico.values.tolist())
            inventarioSA=InventarioHistorico.values.tolist()
            movimiento=[]
            window['-VISUALIZARINVENTARIO-'].update("Disponible")

            costobeneificio()
            figure_canvas_agg.draw()
            window['-REGISTRARENTRADA-'].update(disabled=False)

            try:
                window['-PRODUCTOFOTO-'].update(convToBytes(PIL.Image.open(path+'NoImage.png')))
            except:
                sg.Popup('No existe foto para este producto')
            
            demandagrafico()
            figure_canvas2_agg.draw()
            demandagrafico()
            figure_canvas2_agg.draw()

        try:
            if event == '-REGISTRARENTRADA-':
                if float(values['-CANTIDAD-'])>0 and float(values['-PRECIOVENTA-'])>0 and float(values['-COSTO-'])>0 and values['-CODIGO-'] and values['-PRODUCTO-']:
                
                    nuevoinventario=[values['-FECHA-'],values['-CODIGO-'],values['-PRODUCTO-'],values['-PRECIOVENTA-'],values['-COSTO-'],values['-CANTIDAD-'],float(values['-CANTIDAD-'])*float(values['-PRECIOVENTA-']),float(values['-CANTIDAD-'])*float(values['-COSTO-']),"Entrada"]
                    inventarioSA.append(nuevoinventario)
                    window['-INVENTARIOHISTORICO-'].update(inventarioSA)
                    window['-INVENTARIOHISTORICO-'].set_vscroll_position(1)

                    costobeneificio()
                    figure_canvas_agg.draw()

                    demandagrafico()
                    figure_canvas2_agg.draw() 
                    pd.DataFrame(inventarioSA, columns = ['Fecha','Código', 'Producto', 'Precio Unitario', 'Costo unitario', 'Movimiento', 'Precio Total', 'Costo Total', 'Tipo']).to_csv(path+"InventarioHistorico.csv", index=False, encoding='latin-1')

                    print(foldername)
                    if foldername=="vacio":
                        shutil.copy(path+'noimage.png', path+values['-CODIGO-']+".png")
                    else:
                        shutil.copy(foldername, path+values['-CODIGO-']+".png") 
                else:   
                    sg.Popup('No es una entrada validas')
                
        except:
            sg.Popup('No es una entrada valida')

        if event == '-ELIMINARENTRADA-':
            indexes = values['-INVENTARIOHISTORICO-']
            if indexes:
                for index in sorted(indexes, reverse=True):
                    del inventarioSA[index]
                window['-INVENTARIOHISTORICO-'].update(inventarioSA)
                pd.DataFrame(inventarioSA, columns = ['Fecha','Código', 'Producto', 'Precio Unitario', 'Costo unitario', 'Movimiento', 'Precio Total', 'Costo Total', 'Tipo']).to_csv(path+"InventarioHistorico.csv", index=False, encoding='latin-1')
                costobeneificio()
                figure_canvas_agg.draw()
                demandagrafico()
                figure_canvas2_agg.draw()

        if event == '-INVENTARIOHISTORICO-' and values['-INVENTARIOHISTORICO-']:
            
            data_selected = [inventarioSA[row] for row in values[event]]
            window['-CODIGO-'].update(data_selected[0][1])
            window['-PRODUCTO-'].update(data_selected[0][2])
            window['-COSTO-'].update(data_selected[0][4])
            window['-PRECIOVENTA-'].update(data_selected[0][3])

            window['-CODIGOSALIDA-'].update(data_selected[0][1])
            window['-PRODUCTOSALIDA-'].update(data_selected[0][2])
            window['-PRECIOVENTASALIDA-'].update(data_selected[0][3])

        if event == 'Contabilidad':
            window.close()
            VentanaContable()
        
        if event == '-AGREGARFOTO-':
            if values['-CODIGO-']:
                foldername = sg.popup_get_file('Test', no_window=True, show_hidden=False, file_types=(("Images", "*.png *.gif *.bmp *.jpg *.jpeg"),))
            else:
                sg.Popup('Agregue antes un código de producto')

    window.close()

# %%
def VentanaContable():

        path= os.getcwd()+'\\'

        def verificarfechas():
            expiracion = pd.read_csv(path+'Ajustes.csv',encoding='utf-8')
            expiracion = expiracion.values.tolist()
            expiracion = expiracion[0][0]
            expiracion=datetime.strptime(expiracion, '%b-%d-%Y').date()

            hoy = date.today()

            if hoy<expiracion:
                sg.popup("Demo finalizado")
                sys.exit('listofitems not long enough')

        verificarfechas()

        # Esto es para generar el catalogo de cuentas y el asiento a registrar

        registrodataSA = []
        registrodataSA_heads = ['# Asiento','Fecha','Código', 'Cuenta', 'Debe', 'Haber']
        #registrodataSA.append(['A', 'B', 'C', 'D'])
        registrodataSA_width = [0,0,10, 44, 10, 10]
                
        superiorA= [

                [sg.Text("Fecha", size=(9,1)),sg.Text("Asiento #",size=(9,1)),sg.Text("Nombre de la cuenta")],
                [sg.Input(date.today().strftime("%b-%d-%Y"), size=(12,1), key="-FECHA-", disabled=True),sg.Input("", size=(10,1),key="-#ASIENTO-", disabled=True),sg.Input("", size=(71,None),key="-NOMBRECUENTA-", disabled=True)],
                [sg.Text("Código"),sg.Input("", size=(10,None),key="-CODIGO-", disabled=True),sg.Text("Monto"),sg.Input("", size=(35,None), key="-MONTO-"),sg.Radio('Debe', "RADIO1",key="-DEBE-", default=True),sg.Radio('Haber', "RADIO1",key="-HABER-"),sg.Button("Agregar Cuenta", key="-REGISTRARLINEA-")],
                [sg.Button("Editar Entrada", size=(41,2)),sg.Button("Eliminar Entrada", size=(41,2),key="-ELIMINARENTRADA-", enable_events=True)],
                [sg.Table(values=registrodataSA, headings=registrodataSA_heads,max_col_width=65,col_widths=registrodataSA_width,auto_size_columns=False,justification='left',num_rows=7,background_color="White", row_height=24,text_color="black", key="-ASIENTO-", enable_events=True,select_mode=sg.TABLE_SELECT_MODE_BROWSE)],
                [sg.Button("Registrar Asiento ", size=(84,2), key="-REGISTRARASIENTO-"),]

        ]
        

        #url = 'https://raw.githubusercontent.com/jchavesmartinez/software-financiero/main/cuentas%20-%20Copy.csv'
        totalcuentas = pd.read_csv(path+'cuentas - Copy.csv', encoding='utf-8')
        totalcuentas=totalcuentas.drop(['Nivel', 'Primer Nivel'], axis=1)

        registrodataSB=totalcuentas
        registrodataSB_heads = ['Código', 'Cuenta']
        registrodataSB=registrodataSB.values.tolist()
        registrodataSB_width = [9, 49]

        superiorB= [

            [sg.Button("Agregar Cuenta", expand_x=True, visible=False),sg.Button("Modificar Cuenta", expand_x=True, visible=False),sg.Button("Eliminar Cuenta", expand_x=True, visible=False)],
            [sg.Text("Buscar por nombre"),sg.Input("",expand_x=True, key="-FILTRARCUENTAS-", enable_events=True)],
            [sg.Table(values=registrodataSB, headings=registrodataSB_heads,max_col_width=65,col_widths=registrodataSB_width,auto_size_columns=False,justification='left',num_rows=11,background_color="White", row_height=25,text_color="black", key="-CUENTAS-", enable_events=True,select_mode=sg.TABLE_SELECT_MODE_BROWSE)],
            [sg.Button("Catalogo Completo", size=(32,2), key="-COMPLETO-"),sg.Button("Catalogo Simplificado", size=(32,2), key="-SIMPLIFICADO-"),]

        ]

        # Esto muestra el historico de todos los asientos registrados asi como la opcion de generar los estados financieros

        url = 'https://raw.githubusercontent.com/jchavesmartinez/software-financiero/main/historicoCSV.csv'
        registrodataIA = pd.read_csv(path+'historicoCSV.csv', encoding='utf-8')
        registrodataIA=registrodataIA.values.tolist()
        registrodataIA_heads = ['Asiento','Fecha','Código', 'Cuenta', 'Debe', 'Haber']
        registrodataIA_width = [6,10,10, 68, 10, 10]

        for x in registrodataIA:
            del x[0]

        inferiorA= [

            [sg.Table(values=registrodataIA, headings=registrodataIA_heads,max_col_width=65,col_widths=registrodataIA_width,auto_size_columns=False,justification='left',num_rows=9,background_color="White", row_height=25,text_color="black", key="-HISTORICO-")],
            [sg.Button("Realizar cierre de mes", size=(129,4),key="-CIERREMENSUAL-")]
            
        ]

        inferiorB= [

            [sg.Text("Exportar")],
            [sg.Button("Libro Mayor",size=(22,2),pad=(0,0), key="-MAYORPDF-")],
            [sg.Button("Libro Diario",size=(22,2),pad=(0,0), key="-DIARIOPDF-")],
            [sg.Button("Situación Financiera",size=(22,2),pad=(0,0), key="-RESULTADOSPDF-")],
            [sg.Button("Balance General",size=(22,2),pad=(0,0), key="-BALANCEPDF-")],
            [sg.Text("",size=(10,6))]

        ]

        # Esto consolida todas las partes de la intefaz en un  solo layout

        menucompleto = [
            [sg.Frame("", [[sg.Frame("", superiorA,border_width=0),sg.Frame("", superiorB,border_width=0)]],border_width=0, pad=(0,0))],
            [sg.Frame("", [[sg.Frame("", inferiorA,border_width=0),sg.Frame("", inferiorB,border_width=0)]],border_width=0, pad=(0,0))],
        ]

        menu_def = [['Menu', ['Prónostico de Ventas','Manejo de Inventarios','Facturación', 'Manejo de Planilla', 'Analísis Financiero', 'Cuentas por pagar y cobrar']],['Info', ['Manual de Uso','Preguntas y Respuestas','Reportar Error']],['Acerca de', ['Quienes somos','Contacto']]]
        layout = [
            [sg.Menu(menu_def, tearoff=False, pad=(200, 1))],
            [sg.Frame("", menucompleto,border_width=0,expand_y=True)]
        ]

        right_click_menu = ['Unused', ['Forecasting App', '!&Click', '&Menu', 'E&xit', 'Properties']]
        window= sg.Window("Modulo Contable",layout,finalize=True,icon=path+'favicon2.ico',right_click_menu=right_click_menu)

        # Esta es la funcion para exportar un Excel a PDF

        def exportarpdf(archivoexcel,direccionexportar):
            excel = client.Dispatch("Excel.Application")

            sheets = excel.Workbooks.Open(archivoexcel)
            work_sheets = sheets.Worksheets[0]

            work_sheets.ExportAsFixedFormat(0,direccionexportar)

            sheets.Close()

        # Esto es para crear el Libro Mayor

        def LibroMayor():
            registrodataIA = pd.read_csv(path+'historicoCSV.csv', encoding='utf-8')
            MayorData= pd.DataFrame(registrodataIA, columns =registrodataIA_heads)
            MayorData=MayorData.replace('-','0')
            MayorData['Debe'] = MayorData['Debe'].astype(float)
            MayorData['Haber'] = MayorData['Haber'].astype(float)
            MayorData=MayorData.drop(['Asiento'], axis=1)

            Mayor = MayorData.groupby(['Código','Cuenta']).sum()
            Mayor = Mayor.reset_index() 

            Mayor['Debe Saldos'] = np.where(Mayor['Código'].astype(str).str.startswith('11') | Mayor['Código'].astype(str).str.startswith('12') | Mayor['Código'].astype(str).str.startswith('5') | Mayor['Código'].astype(str).str.startswith('9'), Mayor['Debe']-Mayor['Haber'], 0)
            Mayor['Haber Saldos'] = np.where(Mayor['Código'].astype(str).str.startswith('21') | Mayor['Código'].astype(str).str.startswith('22') | Mayor['Código'].astype(str).str.startswith('3') | Mayor['Código'].astype(str).str.startswith('4') | Mayor['Código'].astype(str).str.startswith('8'), Mayor['Haber']-Mayor['Debe'], 0)


            Mayor['Debe BG'] = np.where(Mayor['Código'].astype(str).str.startswith('11') | Mayor['Código'].astype(str).str.startswith('12') | Mayor['Código'].astype(str).str.startswith('551') ,Mayor['Debe Saldos'] , 0)
            Mayor['Haber BG'] = np.where(Mayor['Código'].astype(str).str.startswith('21') | Mayor['Código'].astype(str).str.startswith('22') | Mayor['Código'].astype(str).str.startswith('3') ,Mayor['Haber Saldos'] , 0)

            Mayor['Debe ER'] = np.where(Mayor['Código'].astype(str).str.startswith('9') | Mayor['Código'].astype(str).str.startswith('515') | Mayor['Código'].astype(str).str.startswith('531') ,Mayor['Debe Saldos'] , 0)
            Mayor['Haber ER'] = np.where(Mayor['Código'].astype(str).str.startswith('4') & ~Mayor['Código'].astype(str).str.startswith('79') ,Mayor['Haber Saldos'] , 0)

            Mayor['Debe ER2'] = np.where(Mayor['Código'].astype(str).str.startswith('5') ,Mayor['Debe Saldos'] , 0)
            Mayor['Haber ER2'] = np.where(Mayor['Código'].astype(str).str.startswith('4') & ~Mayor['Código'].astype(str).str.startswith('79') ,Mayor['Haber Saldos'] , 0)

            MayorExportar=Mayor.drop(['Debe ER','Haber ER','Debe ER2','Haber ER2','Debe BG','Haber BG'], axis=1)

            return MayorExportar , Mayor

        def LibroMayorFormato():

            MayorExportar, Mayor = LibroMayor()

            MayorExportar['Debe'] = np.where(Mayor['Debe'].astype(str).str.startswith('0') ,"-",Mayor['Debe'])
            MayorExportar['Haber'] = np.where(Mayor['Haber'].astype(str).str.startswith('0') ,"-",Mayor['Haber'])
            MayorExportar['Debe Saldos'] = np.where(Mayor['Debe Saldos'].astype(str).str.startswith('0') ,"-",Mayor['Debe Saldos'])
            MayorExportar['Haber Saldos'] = np.where(Mayor['Haber Saldos'].astype(str).str.startswith('0') ,"-",Mayor['Haber Saldos'])


            MayorExportar.to_excel(path+'Mayor.xlsx',sheet_name='Libro Mayor', index=False)

            wb = load_workbook('Mayor.xlsx')
            sheet = wb["Libro Mayor"]

            column_widths = [14,100,17,17,17,17]
            
            for row in MayorExportar:
                for i, cell in enumerate(row):
                    if len(column_widths) > i:
                        if len(cell) > column_widths[i]:
                            column_widths[i] = len(cell)
                    else:
                        column_widths += [len(cell)]
                
            for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
                sheet.column_dimensions[get_column_letter(i)].width = column_width

            thin = Side(border_style="thin", color="00000000")
            double = Side(border_style="double", color="00000000")

            for c in sheet['A1:F1'][0]:
                c.border = Border(bottom=double)
            
            for rows in sheet.iter_rows(min_row=1, max_row=16, min_col=1, max_col=6):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')
            
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_setup.fitToWidth = True

            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.fitToHeight = False
            sheet.page_setup.fitToWidth = 1
            
            wb.save(path+'Mayor.xlsx')

        # Este es el codigo para crear el libro de diario
        def LibroDiario():
            registrodataIA = pd.read_csv(path+'historicoCSV.csv')
            diario= pd.DataFrame (registrodataIA, columns =registrodataIA_heads)
            diario.loc[diario.duplicated(['Asiento']),['Asiento']]=''

            diario.to_excel(path+'LibroDiario.xlsx', index=False,sheet_name='Libro de Diario')

            wb = load_workbook('LibroDiario.xlsx')
            sheet = wb["Libro de Diario"]

            #for rows in sheet.iter_rows(min_row=1, max_row=400, min_col=1, max_col=7):
            #    for cell in rows:
            #        cell.fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type="solid")
            
            column_widths = [10,14,14,120,16,16,16]
            
            for row in diario:
                for i, cell in enumerate(row):
                    if len(column_widths) > i:
                        if len(cell) > column_widths[i]:
                            column_widths[i] = len(cell)
                    else:
                        column_widths += [len(cell)]
                
            for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
                sheet.column_dimensions[get_column_letter(i)].width = column_width

            thin = Side(border_style="thin", color="00000000")
            double = Side(border_style="double", color="00000000")

            for c in sheet['A1:F2'][0]:
                c.border = Border(bottom=double)
            
            for rows in sheet.iter_rows(min_row=1, max_row=160, min_col=1, max_col=8):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')


            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_setup.fitToWidth = True

            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.fitToHeight = False
            sheet.page_setup.fitToWidth = 1

            wb.save(path+'LibroDiario.xlsx')

        # Esto es el codigo para el balance general

        def BalanceGeneral():
            Balance = []
            Balance.append(['','ACTIVOS',''])
            Balance.append(['','ACTIVOS CORRIENTES',''])
            Balance.append(['Código','Cuenta','Saldo'])

            MayorExportar, Mayor = LibroMayor()

            Mayor=Mayor.drop(['Debe ER','Haber ER','Debe ER2','Haber ER2','Debe','Haber','Debe Saldos','Haber Saldos'], axis=1)
            
            # Activos Corrientes
            AC=Mayor[Mayor['Código'].astype(str).str.startswith("11")]
            AC=AC.drop(['Haber BG'], axis=1)
            for x in range(0, len(AC)):
                Balance.append(AC.values.tolist()[x])
            Balance.append(['Total','',AC["Debe BG"].sum()])

            # Activos no corrientes
            Balance.append(['','ACTIVOS NO CORRIENTES',''])
            Balance.append(['Código','Cuenta','Saldo'])
            ANC=Mayor[Mayor['Código'].astype(str).str.startswith("12")]
            ANC=ANC.drop(['Haber BG'], axis=1)

            for x in range(0, len(ANC)):
                Balance.append(ANC.values.tolist()[x])
            Balance.append(['Total','',ANC["Debe BG"].sum()])

            #------------

            Balance.append(['TOTAL ACTIVOS','',ANC["Debe BG"].sum()+AC["Debe BG"].sum()])

            Balance.append(['','',''])
            Balance.append(['','---------------/---------------',''])
            Balance.append(['','',''])

            # Pasivos corrientes
            Balance.append(['','PASIVOS CORRIENTES',''])
            Balance.append(['Código','Cuenta','Saldo'])
            PC=Mayor[Mayor['Código'].astype(str).str.startswith("21")]
            PC=PC.drop(['Debe BG'], axis=1)
            
        
            for x in range(0, len(PC)):
                Balance.append(PC.values.tolist()[x])
            Balance.append(['Total','',PC["Haber BG"].sum()])

            #------------

            # Pasivos no corrientes
            Balance.append(['','PASIVOS NO CORRIENTES',''])
            Balance.append(['Código','Cuenta','Saldo'])
            PNC=Mayor[Mayor['Código'].astype(str).str.startswith("22")]
            PNC=PNC.drop(['Debe BG'], axis=1)
            
        
            for x in range(0, len(PNC)):
                Balance.append(PNC.values.tolist()[x])
            Balance.append(['Total','',PNC["Haber BG"].sum()])
            Balance.append(['TOTAL PASIVOS','',PNC["Haber BG"].sum()+PC["Haber BG"].sum()])

            #------------

            # Patrimonio
            Balance.append(['','PATRIMONIO',''])
            Balance.append(['Código','Cuenta','Saldo'])
            CS=Mayor[Mayor['Código'].astype(str).str.startswith("3")]
            CS=CS.drop(['Debe BG'], axis=1)
            
        
            for x in range(0, len(CS)):
                Balance.append(CS.values.tolist()[x])
            
            Activos=AC["Debe BG"].sum()+ANC["Debe BG"].sum()
            PasivoPatrimonio=CS["Haber BG"].sum()+PC["Haber BG"].sum()+PNC["Haber BG"].sum()
            
            if Activos>PasivoPatrimonio:
                Resultado=Activos-PasivoPatrimonio
            else:
                Resultado=Activos-PasivoPatrimonio
            

            Balance.append(['Resultado del ejercicio','',Resultado])
            Balance.append(['TOTAL PATRIMONIO','',CS["Haber BG"].sum()+Resultado])
            Balance.append(['TOTAL PASIVO Y PATRIMONIO','',PasivoPatrimonio+Resultado])

            #------------

            Balance = pd.DataFrame (Balance, columns=['','BALANCE GENERAL',''])
            Balance.to_excel(path+'BalanceGeneral.xlsx',sheet_name='Balance General', index=False)

            wb = load_workbook('BalanceGeneral.xlsx')
            sheet = wb["Balance General"]
            
            column_widths = [14,100,14]
            
            for row in Balance:
                for i, cell in enumerate(row):
                    if len(column_widths) > i:
                        if len(cell) > column_widths[i]:
                            column_widths[i] = len(cell)
                    else:
                        column_widths += [len(cell)]
                
            for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
                sheet.column_dimensions[get_column_letter(i)].width = column_width

            double = Side(border_style=None, color="00000000")

            for c in sheet['A1:F2'][0]:
                c.border = Border(bottom=double)

            for rows in sheet.iter_rows(min_row=1, max_row=36, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            sheet['B1'].font = Font(name='Cambria',size=20,bold=True)
            sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

            sheet['B2'].font = Font(name='Cambria',size=18,bold=True)
            sheet['B2'].alignment = Alignment(horizontal='center', vertical='center')

            sheet['B3'].font = Font(name='Cambria',size=16,bold=True)
            sheet['B3'].alignment = Alignment(horizontal='center', vertical='center')

            for rows in sheet.iter_rows(min_row=4, max_row=4, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=4+len(AC)+1, max_row=4+len(AC)+1, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=4+len(AC)+1+1, max_row=4+len(AC)+1+1, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=16, bold=True)
                    cell.alignment  = Alignment(horizontal='center', vertical='center')

            for rows in sheet.iter_rows(min_row=7+len(AC), max_row=7+len(AC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')
            
            for rows in sheet.iter_rows(min_row=8+len(AC)+len(ANC), max_row=8+len(AC)+len(ANC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=9+len(AC)+len(ANC), max_row=9+len(AC)+len(ANC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=11+len(AC)+len(ANC), max_row=11+len(AC)+len(ANC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=14, bold=True)
                    cell.alignment  = Alignment(horizontal='center', vertical='center')

            for rows in sheet.iter_rows(min_row=13+len(AC)+len(ANC), max_row=13+len(AC)+len(ANC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=16, bold=True)
                    cell.alignment  = Alignment(horizontal='center', vertical='center')
            
            for rows in sheet.iter_rows(min_row=14+len(AC)+len(ANC), max_row=14+len(AC)+len(ANC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=15+len(AC)+len(ANC)+len(PC), max_row=15+len(AC)+len(ANC)+len(PC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=16+len(AC)+len(ANC)+len(PC), max_row=16+len(AC)+len(ANC)+len(PC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=16, bold=True)
                    cell.alignment  = Alignment(horizontal='center', vertical='center')

            for rows in sheet.iter_rows(min_row=17+len(AC)+len(ANC)+len(PC), max_row=17+len(AC)+len(ANC)+len(PC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')
            
            for rows in sheet.iter_rows(min_row=18+len(AC)+len(ANC)+len(PC)+len(PNC), max_row=18+len(AC)+len(ANC)+len(PC)+len(PNC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')
            
            for rows in sheet.iter_rows(min_row=19+len(AC)+len(ANC)+len(PC)+len(PNC), max_row=19+len(AC)+len(ANC)+len(PC)+len(PNC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=20+len(AC)+len(ANC)+len(PC)+len(PNC), max_row=20+len(AC)+len(ANC)+len(PC)+len(PNC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=16, bold=True)
                    cell.alignment  = Alignment(horizontal='center', vertical='center')
            
            for rows in sheet.iter_rows(min_row=21+len(AC)+len(ANC)+len(PC)+len(PNC), max_row=21+len(AC)+len(ANC)+len(PC)+len(PNC), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=22+len(AC)+len(ANC)+len(PC)+len(PNC)+len(CS), max_row=22+len(AC)+len(ANC)+len(PC)+len(PNC)+len(CS), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, italic=True )
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=23+len(AC)+len(ANC)+len(PC)+len(PNC)+len(CS), max_row=23+len(AC)+len(ANC)+len(PC)+len(PNC)+len(CS), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')\

            for rows in sheet.iter_rows(min_row=24+len(AC)+len(ANC)+len(PC)+len(PNC)+len(CS), max_row=24+len(AC)+len(ANC)+len(PC)+len(PNC)+len(CS), min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_setup.fitToWidth = True

            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.fitToHeight = False
            sheet.page_setup.fitToWidth = 1  

            wb.save(path+'BalanceGeneral.xlsx')
        
        def EstadoResultados():
            Situacion = []
            Situacion.append(['','ESTADO DE PÉRDIDAS Y GANANCIAS',''])
            Situacion.append(['INGRESOS OPERATIVOS','',''])
            Situacion.append(['Código','Cuenta','Saldo'])

            MayorExportar, Mayor = LibroMayor()

            # Ingresos
            Mayor1=Mayor.drop(['Debe ER2','Haber ER2','Debe','Haber','Debe BG','Haber BG','Debe Saldos','Haber Saldos'], axis=1)
            Ingresos=Mayor1[Mayor1['Código'].astype(str).str.startswith("44") | Mayor1['Código'].astype(str).str.startswith("5999902") | Mayor1['Código'].astype(str).str.startswith("4999902") | Mayor1['Código'].astype(str).str.startswith("531")]
            Ingresos["SaldosNetos"] = Ingresos["Debe ER"]+Ingresos["Haber ER"]
            Ingresos=Ingresos.drop(['Debe ER','Haber ER',], axis=1)
            Ingresos = Ingresos[Ingresos['SaldosNetos'] != 0]
            for x in range(0, len(Ingresos)):
                Situacion.append(Ingresos.values.tolist()[x])
            rentabruta=Ingresos["SaldosNetos"].sum()
            Situacion.append(['UTILIDAD BRUTA','',rentabruta])

            # GASTOS OPERATIVOS
            Situacion.append(['GASTOS OPERATIVOS','',''])
            Mayor2=Mayor.drop(['Debe ER','Haber ER','Debe','Haber','Debe BG','Haber BG','Debe Saldos','Haber Saldos'], axis=1)
            GastosOperativos=Mayor2[Mayor2['Código'].astype(str).str.startswith("51")]
            GastosOperativos["SaldosNetos"] = GastosOperativos["Debe ER2"]+GastosOperativos["Haber ER2"]
            GastosOperativos=GastosOperativos.drop(['Debe ER2','Haber ER2',], axis=1)
            GastosOperativos = GastosOperativos[GastosOperativos['SaldosNetos'] != 0]
            for x in range(0, len(GastosOperativos)):
                Situacion.append(GastosOperativos.values.tolist()[x])
            Situacion.append(['UTILIDAD OPERATIVA','',Ingresos["SaldosNetos"].sum()-GastosOperativos["SaldosNetos"].sum()])


            # GASTOS FINANCIEROS
            Situacion.append(['GASTOS FINANCIEROS','',''])
            GastosFinancieros=Mayor2[Mayor2['Código'].astype(str).str.startswith("52")]
            GastosFinancieros["SaldosNetos"] = GastosFinancieros["Debe ER2"]+GastosFinancieros["Haber ER2"]
            GastosFinancieros=GastosFinancieros.drop(['Debe ER2','Haber ER2',], axis=1)
            GastosFinancieros = GastosFinancieros[GastosFinancieros['SaldosNetos'] != 0]
            for x in range(0, len(GastosFinancieros)):
                Situacion.append(GastosFinancieros.values.tolist()[x])
            #Situacion.append(['UTILIDAD ANTES DE IMPUESTOS','',Ingresos["SaldosNetos"].sum()-GastosOperativos["SaldosNetos"].sum()-GastosFinancieros["SaldosNetos"].sum()])

            # OTROS INGRESOS
            Situacion.append(['OTROS INGRESOS','',''])
            OtrosIngresos=Mayor2[Mayor2['Código'].astype(str).str.startswith("41") | Mayor2['Código'].astype(str).str.startswith("42") | Mayor2['Código'].astype(str).str.startswith("43") | Mayor2['Código'].astype(str).str.startswith("45") | Mayor2['Código'].astype(str).str.startswith("46") | Mayor2['Código'].astype(str).str.startswith("49")]
            OtrosIngresos["SaldosNetos"] = OtrosIngresos["Debe ER2"]+OtrosIngresos["Haber ER2"]
            OtrosIngresos=OtrosIngresos.drop(['Debe ER2','Haber ER2',], axis=1)
            OtrosIngresos = OtrosIngresos[OtrosIngresos['SaldosNetos'] != 0]
            for x in range(0, len(OtrosIngresos)):
                Situacion.append(OtrosIngresos.values.tolist()[x])
            #Situacion.append(['UTILIDAD OPERATIVA3','',Ingresos["SaldosNetos"].sum()-GastosOperativos["SaldosNetos"].sum()-GastosFinancieros["SaldosNetos"].sum()+OtrosIngresos["SaldosNetos"].sum()])

            # OTROS GASTOS
            Situacion.append(['OTROS GASTOS','',''])
            OtrosGastos=Mayor2[Mayor2['Código'].astype(str).str.startswith("54") | Mayor2['Código'].astype(str).str.startswith("55") | Mayor2['Código'].astype(str).str.startswith("59")]
            OtrosGastos["SaldosNetos"] = OtrosGastos["Debe ER2"]+OtrosGastos["Haber ER2"]
            OtrosGastos=OtrosGastos.drop(['Debe ER2','Haber ER2',], axis=1)
            OtrosGastos = OtrosGastos[OtrosGastos['SaldosNetos'] != 0]
            for x in range(0, len(OtrosGastos)):
                Situacion.append(OtrosGastos.values.tolist()[x])
            rentaneta=Ingresos["SaldosNetos"].sum()-GastosOperativos["SaldosNetos"].sum()-OtrosIngresos["SaldosNetos"].sum()-GastosFinancieros["SaldosNetos"].sum()+OtrosIngresos["SaldosNetos"].sum()-OtrosGastos["SaldosNetos"].sum()
            Situacion.append(['UTILIDAD ANTES DE IMPUESTOS','',rentaneta])

            if rentabruta>112070000:
                impuesto=rentaneta*0.3
                   #print("Renta bruta mayor a 112.070.000")
            if rentabruta<=112070000:
                if rentaneta<=5286000:
                    impuesto=rentaneta*0.05
                    #print("rentaneta<=5286000:")
                if 5286000 < rentaneta <= 7930000:
                    impuesto=((rentaneta-5286000)*0.1)+5286000*0.05
                    #print("5286000 < rentaneta <= 7930000")
                if 7930000 < rentaneta <= 10573000:
                    impuesto=((rentaneta-7930000)*0.15)+5286000*0.05+((7930000-5286000)*0.10)
                    #print("7930000 < rentaneta <= 10573000")
                if rentaneta>= 10573000:
                    impuesto=((rentaneta-7930000)*0.15)+5286000*0.05+((7930000-5286000)*0.10)+((10573000-7930000)*0.15)
                    #print("rentaneta>= 10573000:")

            Situacion.append(['Impuesto a la renta','',impuesto])
            Situacion.append(['UTILIDAD NETA','',rentaneta-impuesto])

            Situacion = pd.DataFrame (Situacion, columns=['','ESTADO DE RESULTADOS INTEGRAL',''])
            Situacion.to_excel(path+'EstadoResultados.xlsx',sheet_name='Estado de Resultados', index=False)
            Mayor.to_excel(path+'MAYOR2.xlsx',sheet_name='Estado de Resultados', index=False)

            wb = load_workbook('EstadoResultados.xlsx')
            sheet = wb["Estado de Resultados"]
            
            column_widths = [14,100,14]
            
            for row in Situacion:
                for i, cell in enumerate(row):
                    if len(column_widths) > i:
                        if len(cell) > column_widths[i]:
                            column_widths[i] = len(cell)
                    else:
                        column_widths += [len(cell)]
                
            for i, column_width in enumerate(column_widths,1):  # ,1 to start at 1
                sheet.column_dimensions[get_column_letter(i)].width = column_width

            double = Side(border_style=None, color="00000000")

            for c in sheet['A1:F2'][0]:
                c.border = Border(bottom=double)

            for rows in sheet.iter_rows(min_row=1, max_row=36, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            sheet['B1'].font = Font(name='Cambria',size=20,bold=True)
            sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

            sheet['B2'].font = Font(name='Cambria',size=18,bold=True)
            sheet['B2'].alignment = Alignment(horizontal='center', vertical='center')

            sheet['A3'].font = Font(name='Cambria',size=16,bold=True)
            sheet['A3'].alignment = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=4, max_row=4, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=4+len(Ingresos)+1, max_row=4+len(Ingresos)+1, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=4+len(Ingresos)+1+1, max_row=4+len(Ingresos)+1+1, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=16, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=len(Ingresos)+len(GastosOperativos)+7, max_row=len(Ingresos)+len(GastosOperativos)+7, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=len(Ingresos)+len(GastosOperativos)+8, max_row=len(Ingresos)+len(GastosOperativos)+8, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=16, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=len(Ingresos)+len(GastosOperativos)+len(GastosFinancieros)+9, max_row=len(Ingresos)+len(GastosOperativos)+len(GastosFinancieros)+9, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=16, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=len(Ingresos)+len(GastosOperativos)+len(GastosFinancieros)+len(OtrosIngresos)+10, max_row=len(Ingresos)+len(GastosOperativos)+len(GastosFinancieros)+len(OtrosIngresos)+10, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=16, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=len(Ingresos)+len(GastosOperativos)+len(GastosFinancieros)+len(OtrosIngresos)+len(OtrosGastos)+11, max_row=len(Ingresos)+len(GastosOperativos)+len(GastosFinancieros)+len(OtrosIngresos)+len(OtrosGastos)+11, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=len(Ingresos)+len(GastosOperativos)+len(GastosFinancieros)+len(OtrosIngresos)+len(OtrosGastos)+12, max_row=len(Ingresos)+len(GastosOperativos)+len(GastosFinancieros)+len(OtrosIngresos)+len(OtrosGastos)+12, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, italic=True )
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            for rows in sheet.iter_rows(min_row=len(Ingresos)+len(GastosOperativos)+len(GastosFinancieros)+len(OtrosIngresos)+len(OtrosGastos)+13, max_row=len(Ingresos)+len(GastosOperativos)+len(GastosFinancieros)+len(OtrosIngresos)+len(OtrosGastos)+13, min_col=1, max_col=3):
                for cell in rows:
                    cell.font = Font(name='Cambria',size=12, bold=True)
                    cell.alignment  = Alignment(horizontal='left', vertical='center')

            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_setup.fitToWidth = True

            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.fitToHeight = False
            sheet.page_setup.fitToWidth = 1  

            wb.save(path+'EstadoResultados.xlsx')
                

        # Esto es el codigo para ejecutar la interfaz


        while True:
            
            event, values = window.read()

            
            if event == sg.WIN_CLOSED:
                break
            
            
            if event == "-ELIMINARENTRADA-":
                indexes = values['-ASIENTO-']
                if indexes:
                    for index in sorted(indexes, reverse=True):
                        del registrodataSA[index]
                    window['-ASIENTO-'].update(registrodataSA)
 
            if event == 'Manejo de Inventarios':
                window.close()
                VentanaInventarios()
            
            if event == '-CUENTAS-'and values['-CUENTAS-']:
                data_selected = [registrodataSB[row] for row in values[event]]
                window['-CODIGO-'].update(data_selected[0][0])
                window['-NOMBRECUENTA-'].update(data_selected[0][1])
                if len(registrodataIA)==0:
                    numero=1
                else:
                    numero=int(registrodataIA[len(registrodataIA)-1][0])+1
                window['-#ASIENTO-'].update(numero)

            if event == '-SIMPLIFICADO-':
                window['-FILTRARCUENTAS-'].update("")
                cuentasimplificado=totalcuentas.loc[(totalcuentas['Código'] <= 99999)].values.tolist()
                window['-CUENTAS-'].update(cuentasimplificado)
                registrodataSB=cuentasimplificado
            if event == '-COMPLETO-':
                cuentatotal=totalcuentas.loc[(totalcuentas['Código'] <= 9999999999)].values.tolist()
                window['-CUENTAS-'].update(cuentatotal)
                registrodataSB=cuentatotal
            if event == '-FILTRARCUENTAS-' and values['-FILTRARCUENTAS-']:
                filtrarcuentas=totalcuentas.applymap(str)
                filtro=filtrarcuentas[filtrarcuentas.stack().str.contains(values['-FILTRARCUENTAS-']).any(level=0)].values.tolist()
                window['-CUENTAS-'].update(filtro)
                registrodataSB=filtro
            elif event == '-FILTRARCUENTAS-' and values['-FILTRARCUENTAS-']=="":
                cuentatotal=totalcuentas.loc[(totalcuentas['Código'] <= 9999999999)].values.tolist()
                window['-CUENTAS-'].update(cuentatotal)
                registrodataSB=cuentatotal
            try:
                if event == '-REGISTRARLINEA-' and float(values['-MONTO-'])>0 and values['-NOMBRECUENTA-'] and values['-CODIGO-'] :
                    nuevoasiento=[values['-#ASIENTO-'],values['-FECHA-'],data_selected[0][0],data_selected[0][1], values['-MONTO-'] if values['-DEBE-']==True else "-", values['-MONTO-'] if values['-HABER-']==True else "-"]
                    registrodataSA.append(nuevoasiento)
                    window['-ASIENTO-'].update(registrodataSA)
                    window['-ASIENTO-'].set_vscroll_position(1)
            except:
                sg.Popup('No es un asiento válido, por favor revise los campos')
            
            try:              
                if event == '-REGISTRARASIENTO-':        
                    debe=[]
                    haber=[]

                    for x in range(0, len(registrodataSA)):
                        debe.append(registrodataSA[x][4])
                        haber.append(registrodataSA[x][5])

                    debe = [s.replace("-", '0') for s in debe]
                    haber = [s.replace("-", '0') for s in haber]

                    validacion=sum(map(float, debe))==sum(map(float, haber))

                    if validacion==True:
                        for x in range(0, len(registrodataSA)):
                            registrodataIA.append(registrodataSA[x])
                        window['-HISTORICO-'].update(registrodataIA)
                        window['-NOMBRECUENTA-'].update("")
                        window['-CODIGO-'].update("")
                        window['-MONTO-'].update("")
                        registrodataSA = []
                        window['-ASIENTO-'].update(registrodataSA)
                        if len(registrodataIA)==0:
                            numero=1
                        else:
                            numero=int(registrodataIA[len(registrodataIA)-1][0])+1
                        window['-#ASIENTO-'].update(numero)
                        window['-FECHA-'].update(date.today().strftime("%b-%d-%Y"))
                        window['-HISTORICO-'].set_vscroll_position(1)

                        pd.DataFrame(registrodataIA, columns = registrodataIA_heads).to_csv("historicoCSV.csv")
                    else:
                        sg.Popup('No es un asiento válido, por favor revise los campos')

            except:
                sg.Popup('No es un asiento válido, por favor revise los campos') 

            if event == '-DIARIOPDF-':

                LibroDiario()
                archivoexcel= path+'LibroDiario.xlsx'
                direccionexportar= path+'LibroDiario.pdf'
                exportarpdf(archivoexcel,direccionexportar)
                sg.popup("Libro de diario generado exitosamente")


            if event == '-MAYORPDF-':
                
                LibroMayor()
                LibroMayorFormato()
                archivoexcel= path+'Mayor.xlsx'
                direccionexportar= path+'Mayor.pdf'
                exportarpdf(archivoexcel,direccionexportar)
                sg.popup("Libro mayor generado exitosamente")

            if event == '-RESULTADOSPDF-':  
                EstadoResultados()
                archivoexcel= path+'EstadoResultados.xlsx'
                direccionexportar= path+'EstadoResultados.pdf'
                exportarpdf(archivoexcel,direccionexportar)
                sg.popup("Libro de diario generado exitosamente")


            if event == '-BALANCEPDF-':
                
                BalanceGeneral()
                archivoexcel= path+'BalanceGeneral.xlsx'
                direccionexportar= path+'BalanceGeneral.pdf'
                exportarpdf(archivoexcel,direccionexportar)
                sg.popup("Balance general creado exitosamente")
    
        
            if event == "-CIERREMENSUAL-":
                # Cierre de gastos
                MayorExportar, Mayor = LibroMayor()
                MayorCierre=Mayor.drop(['Debe ER','Haber ER','Debe','Haber','Debe BG','Haber BG','Debe Saldos','Haber Saldos'], axis=1)
                GastosCierre=MayorCierre[MayorCierre['Código'].astype(str).str.startswith("51") | MayorCierre['Código'].astype(str).str.startswith("52") | MayorCierre['Código'].astype(str).str.startswith("53") | MayorCierre['Código'].astype(str).str.startswith("54") | MayorCierre['Código'].astype(str).str.startswith("59")]
                GastosCierre['Haber ER2'] = GastosCierre['Debe ER2']
                GastosCierre['Debe ER2'] = 0
                GastosCierre.insert(0, 'Fecha', date.today().strftime("%b-%d-%Y"))
                GastosCierre.insert(0, 'Asiento', int(registrodataIA[len(registrodataIA)-1][0])+1)
                registrodataIA.append([int(registrodataIA[len(registrodataIA)-1][0])+1,date.today().strftime("%b-%d-%Y"),"31502","Resultado del ejercicio",GastosCierre['Haber ER2'].sum(),0])
                for x in range(0, len(GastosCierre)):
                    registrodataIA.append(GastosCierre.values.tolist()[x])
                window['-HISTORICO-'].update(registrodataIA)
                pd.DataFrame(registrodataIA, columns = registrodataIA_heads).to_csv("historicoCSV.csv")


        print(registrodataIA)


        window.close()

VentanaContable()


# %%

